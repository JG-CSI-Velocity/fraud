#!/usr/bin/env python3
"""Full fraud screening pipeline.

Usage:
    python run.py clients/acme/          # new client, interactive setup
    python run.py clients/acme/ --rerun  # re-run with existing config

Workflow:
    1. Finds the Excel/CSV data file in the client folder
    2. If no YAML config exists, auto-detects column mappings
    3. Runs the fraud screener
    4. Saves the report to the client folder
"""

import argparse
import subprocess
import sys
from pathlib import Path

import yaml

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# Keywords used to auto-detect which spreadsheet column maps to each field.
# Each field has a list of patterns checked against the lowercase header.
# More specific patterns go first so they match before generic ones.
AUTO_MATCH = {
    "manager": ["purchase manager", "manager", "employee", "staff", "rep name"],
    "account_holder": ["account holder", "account name", "new member", "member name", "customer"],
    "referrer": ["referrer", "referred by", "referral name", "referring"],
    "row_number": ["row", "#"],
    "branch_number": ["branch number", "branch num", "branch id", "branch #"],
    "branch_name": ["branch name", "branch"],
    "issue_date": ["issue date", "date", "created"],
    "certificate_id": ["certificate", "cert id", "tracking id", "cert"],
    "referral_code": ["referral code", "ref code", "code"],
    "program_name": ["program name", "program"],
    "product_count": ["product count", "products", "# products"],
}

REQUIRED_FIELDS = ["manager", "account_holder", "referrer"]

FIELD_DESCRIPTIONS = {
    "manager": "Employee who set up the referral",
    "account_holder": "Person who was referred",
    "referrer": "Person who referred them",
    "row_number": "Row number or ID",
    "branch_number": "Branch ID or number",
    "branch_name": "Branch name",
    "issue_date": "Date the referral was issued",
    "certificate_id": "Certificate or tracking ID",
    "referral_code": "Referral code",
    "program_name": "Program name",
    "product_count": "Number of products opened",
}


def find_data_file(folder: Path) -> Path | None:
    candidates = [
        f for f in folder.iterdir()
        if f.is_file()
        and f.suffix.lower() in SUPPORTED_EXTENSIONS
        and "_fraud_report" not in f.stem.lower()
    ]
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    print("\nMultiple data files found:")
    for i, f in enumerate(candidates, 1):
        print(f"  {i}. {f.name}")
    print(f"  Select file (1-{len(candidates)}): ", end="")
    while True:
        choice = input().strip()
        try:
            idx = int(choice)
            if 1 <= idx <= len(candidates):
                return candidates[idx - 1]
        except ValueError:
            pass
        print(f"  Enter 1-{len(candidates)}: ", end="")


def find_config(folder: Path) -> Path | None:
    configs = [f for f in folder.iterdir() if f.suffix.lower() in (".yaml", ".yml")]
    if configs:
        return configs[0]
    return None


def read_headers(data_file: Path):
    """Read column headers and return (headers, sheet_name)."""
    import csv
    import openpyxl

    sheet_name = None
    if data_file.suffix.lower() in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(data_file, data_only=True, read_only=True)
        ws = wb.active
        sheet_name = ws.title
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        sheets = wb.sheetnames
        wb.close()

        if len(sheets) > 1:
            print(f"\nSheets found: {', '.join(sheets)}")
            print(f"  Using: {sheet_name}")
            print(f"  Change sheet? Enter name or press Enter to keep: ", end="")
            choice = input().strip()
            if choice and choice in sheets:
                sheet_name = choice
                wb = openpyxl.load_workbook(data_file, data_only=True, read_only=True)
                ws = wb[sheet_name]
                headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                wb.close()
    else:
        with open(data_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)

    return headers, sheet_name


def auto_match_columns(headers: list[str]) -> dict[str, str]:
    """Try to automatically match headers to screener fields."""
    matches = {}
    used_headers = set()
    header_lower = [(h or "").strip().lower() for h in headers]

    for field, patterns in AUTO_MATCH.items():
        for pattern in patterns:
            for i, h in enumerate(header_lower):
                if h and h not in used_headers and pattern in h:
                    matches[field] = headers[i]
                    used_headers.add(h)
                    break
            if field in matches:
                break

    return matches


def prompt_manual_mapping(headers: list[str], field: str, required: bool) -> str | None:
    """Ask user to pick a column number for a field."""
    label = "REQUIRED" if required else "optional"
    desc = FIELD_DESCRIPTIONS[field]
    print(f"  [{label}] {field} ({desc})")
    print(f"    Enter column number (1-{len(headers)}), or Enter to skip: ", end="")

    while True:
        choice = input().strip()
        if not choice:
            if required:
                print(f"    This field is required. Enter a column number: ", end="")
                continue
            return None
        try:
            idx = int(choice)
            if 1 <= idx <= len(headers):
                return headers[idx - 1]
            else:
                print(f"    Enter 1-{len(headers)}: ", end="")
        except ValueError:
            print(f"    Enter a number or press Enter to skip: ", end="")


def create_config(data_file: Path, folder: Path) -> Path:
    """Auto-detect columns, show results, let user confirm or fix."""
    headers, sheet_name = read_headers(data_file)

    # Show columns
    print(f"\nColumns in {data_file.name}:")
    print("-" * 50)
    for i, h in enumerate(headers, 1):
        print(f"  {i:3d}. {h if h else '(empty)'}")

    # Auto-match
    matches = auto_match_columns(headers)

    # Show auto-detected mappings
    print("\n" + "=" * 50)
    print("AUTO-DETECTED COLUMN MAPPING")
    print("=" * 50)

    all_fields = list(FIELD_DESCRIPTIONS.keys())
    matched_fields = []
    unmatched_required = []
    unmatched_optional = []

    for field in all_fields:
        req = field in REQUIRED_FIELDS
        desc = FIELD_DESCRIPTIONS[field]
        if field in matches:
            tag = "REQUIRED" if req else "optional"
            print(f"  [{tag}] {field} -> \"{matches[field]}\"")
            matched_fields.append(field)
        elif req:
            unmatched_required.append(field)
        else:
            unmatched_optional.append(field)

    if unmatched_required:
        print(f"\n  Could not auto-detect: {', '.join(unmatched_required)}")
    if unmatched_optional:
        print(f"  Not matched (optional): {', '.join(unmatched_optional)}")

    # Confirm or fix
    if unmatched_required:
        print("\nSome required fields need manual mapping:")
        for field in unmatched_required:
            result = prompt_manual_mapping(headers, field, required=True)
            if result:
                matches[field] = result

    print(f"\nAccept this mapping? [Y/n]: ", end="")
    choice = input().strip().lower()

    if choice in ("n", "no"):
        print("\nFix mappings by entering a column number, or Enter to keep:")
        for field in all_fields:
            current = matches.get(field)
            desc = FIELD_DESCRIPTIONS[field]
            req = field in REQUIRED_FIELDS
            tag = "REQUIRED" if req else "optional"
            current_display = f" (current: \"{current}\")" if current else " (not set)"
            print(f"  [{tag}] {field}{current_display}: ", end="")

            while True:
                val = input().strip()
                if not val:
                    break
                try:
                    idx = int(val)
                    if 1 <= idx <= len(headers):
                        matches[field] = headers[idx - 1]
                        print(f"    -> \"{headers[idx - 1]}\"")
                        break
                    else:
                        print(f"    Enter 1-{len(headers)} or Enter to keep: ", end="")
                except ValueError:
                    print(f"    Enter a number or Enter to keep: ", end="")

    # Collect client info
    client_guess = folder.name.replace("-", " ").replace("_", " ").title()
    print(f"\nClient name [{client_guess}]: ", end="")
    client_input = input().strip()
    client_name = client_input if client_input else client_guess

    print("Program name: ", end="")
    program_name = input().strip()

    print("Dollar value per referral [25.00]: ", end="")
    dollar_input = input().strip()
    dollar_value = float(dollar_input) if dollar_input else 25.00

    # Build config
    columns = {k: v for k, v in matches.items() if v}
    config = {
        "client": client_name,
        "program": program_name,
        "columns": columns,
        "dollar_value_per_referral": dollar_value,
    }
    if sheet_name:
        config["sheet"] = sheet_name

    # Final confirmation
    print(f"\n{'=' * 50}")
    print("FINAL CONFIG")
    print("=" * 50)
    print(f"  Client:   {client_name}")
    print(f"  Program:  {program_name}")
    print(f"  $/Referral: {dollar_value}")
    print(f"\n  Column mappings:")
    for field, header in columns.items():
        req = "*" if field in REQUIRED_FIELDS else " "
        print(f"   {req} {field:20s} -> \"{header}\"")
    print(f"\n  * = required")

    print(f"\nSave and run? [Y/n]: ", end="")
    confirm = input().strip().lower()
    if confirm in ("n", "no"):
        print("Aborted.")
        sys.exit(0)

    config_path = folder / "config.yaml"
    with open(config_path, "w") as f:
        yaml.dump(config, f, default_flow_style=False, sort_keys=False)
    print(f"Config saved to {config_path}")

    return config_path


def main():
    parser = argparse.ArgumentParser(description="Full fraud screening pipeline")
    parser.add_argument("folder", help="Path to client folder containing data file")
    parser.add_argument("--rerun", action="store_true", help="Re-run with existing config")
    parser.add_argument("--quiet", "-q", action="store_true", help="Suppress screener progress")
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.is_dir():
        print(f"Error: {folder} is not a directory")
        sys.exit(1)

    # Step 1: Find data file
    data_file = find_data_file(folder)
    if not data_file:
        print(f"Error: No Excel/CSV file found in {folder}")
        sys.exit(1)
    print(f"Data file: {data_file.name}")

    # Step 2: Find or create config
    config_path = find_config(folder)
    if config_path and args.rerun:
        print(f"Config found: {config_path.name}")
    elif config_path:
        print(f"Config found: {config_path.name}")
        print(f"  Use existing config? [Y/n]: ", end="")
        choice = input().strip().lower()
        if choice in ("n", "no"):
            config_path = None

    if not config_path:
        config_path = create_config(data_file, folder)

    # Step 3: Run screener
    output_name = f"{data_file.stem}_fraud_report.xlsx"
    output_path = folder / output_name

    print(f"\nRunning fraud screener...")
    cmd = [
        sys.executable, "screen.py",
        str(data_file),
        "--config", str(config_path),
        "--output", str(output_path),
    ]
    if args.quiet:
        cmd.append("--quiet")

    result = subprocess.run(cmd, cwd=Path(__file__).parent)

    if result.returncode != 0:
        print(f"\nScreener failed with exit code {result.returncode}")
        sys.exit(result.returncode)

    print(f"\nReport saved to: {output_path}")


if __name__ == "__main__":
    main()
