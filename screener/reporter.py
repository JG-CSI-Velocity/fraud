from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .rules.base import Finding

TIER_FILLS = {
    "CRITICAL": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "HIGH": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "LOW": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
}

SEVERITY_FILLS = {
    "critical": TIER_FILLS["CRITICAL"],
    "high": TIER_FILLS["HIGH"],
    "medium": TIER_FILLS["MEDIUM"],
    "low": TIER_FILLS["LOW"],
}

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")


def generate_report(
    findings: list[Finding],
    scores: dict,
    config: dict,
    records: list[dict],
    output_path: str,
):
    wb = Workbook()
    _write_executive_summary(wb.active, findings, scores, config, records)
    _write_referrer_scorecard(wb.create_sheet("Referrer Scorecard"), scores, config)
    _write_manager_scorecard(wb.create_sheet("Manager Scorecard"), scores, config, records)
    _write_branch_scorecard(wb.create_sheet("Branch Scorecard"), scores, config)
    _write_flagged_records(wb.create_sheet("Flagged Records"), findings)
    _write_rule_findings(wb.create_sheet("Self-Referrals"), findings, "Self-Referral")
    _write_rule_findings(wb.create_sheet("Batch Events"), findings, "Batch Referral")
    _write_rule_findings(wb.create_sheet("Rings & Pairs"), findings, "Ring Detection", "Reciprocal Pair")
    _write_rule_findings(wb.create_sheet("Duplicates"), findings, "Duplicate Account")
    _write_rule_findings(wb.create_sheet("Data Quality"), findings, "Data Quality")
    _write_config_tab(wb.create_sheet("Configuration"), config)

    wb.save(output_path)


def _add_header_row(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _write_executive_summary(ws, findings, scores, config, records):
    ws.title = "Executive Summary"
    client = config.get("client", "Unknown")
    dollar = config.get("dollar_value_per_referral", 25)

    total_records = len(records)
    total_findings = len(findings)
    flagged_rows = set()
    for f in findings:
        flagged_rows.update(f.row_numbers)

    rule_counts = defaultdict(int)
    for f in findings:
        rule_counts[f.rule_name] += 1

    severity_counts = defaultdict(int)
    for f in findings:
        severity_counts[f.severity] += 1

    ws.cell(row=1, column=1, value=f"Referral Fraud Screening Report - {client}").font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=3, column=1, value=f"Program: {config.get('program', 'N/A')}")

    row = 5
    ws.cell(row=row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Value").font = HEADER_FONT

    metrics = [
        ("Total Records Screened", total_records),
        ("Total Findings", total_findings),
        ("Unique Rows Flagged", len(flagged_rows)),
        ("Flag Rate", f"{len(flagged_rows)/total_records*100:.1f}%" if total_records else "0%"),
        ("Estimated Exposure (flagged rows)", f"${len(flagged_rows) * dollar:,.0f}"),
        ("", ""),
        ("Critical Findings", severity_counts.get("critical", 0)),
        ("High Findings", severity_counts.get("high", 0)),
        ("Medium Findings", severity_counts.get("medium", 0)),
        ("Low Findings", severity_counts.get("low", 0)),
    ]

    for i, (label, value) in enumerate(metrics, 1):
        ws.cell(row=row + i, column=1, value=label)
        ws.cell(row=row + i, column=2, value=value)

    row = row + len(metrics) + 3
    ws.cell(row=row, column=1, value="Findings by Rule").font = HEADER_FONT
    row += 1
    for rule, count in sorted(rule_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=rule)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Top 15 Referrers by Risk Score").font = HEADER_FONT
    row += 1
    _add_header_row_at(ws, row, ["Referrer", "Score", "Tier", "Rules Hit"])
    row += 1
    ref_scores = scores.get("referrer", {})
    top_refs = sorted(ref_scores.items(), key=lambda x: -x[1]["score"])[:15]
    for ref, data in top_refs:
        ws.cell(row=row, column=1, value=ref)
        ws.cell(row=row, column=2, value=data["score"])
        tier_cell = ws.cell(row=row, column=3, value=data["tier"])
        tier_cell.fill = TIER_FILLS.get(data["tier"], PatternFill())
        ws.cell(row=row, column=4, value=", ".join(sorted(data["rules_hit"])))
        row += 1

    _auto_width(ws)


def _add_header_row_at(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL


def _write_referrer_scorecard(ws, scores, config):
    dollar = config.get("dollar_value_per_referral", 25)
    headers = ["Referrer", "Score", "Tier", "Rules Hit", "Finding Count", "Exposure"]
    _add_header_row(ws, headers)

    ref_scores = scores.get("referrer", {})
    for row_idx, (ref, data) in enumerate(
        sorted(ref_scores.items(), key=lambda x: -x[1]["score"]), 2
    ):
        ws.cell(row=row_idx, column=1, value=ref)
        ws.cell(row=row_idx, column=2, value=data["score"])
        tier_cell = ws.cell(row=row_idx, column=3, value=data["tier"])
        tier_cell.fill = TIER_FILLS.get(data["tier"], PatternFill())
        ws.cell(row=row_idx, column=4, value=", ".join(sorted(data["rules_hit"])))
        finding_count = len(data["findings"])
        ws.cell(row=row_idx, column=5, value=finding_count)
        ws.cell(row=row_idx, column=6, value=f"${finding_count * dollar:,.0f}")

    _auto_width(ws)


def _write_manager_scorecard(ws, scores, config, records):
    from collections import Counter

    dollar = config.get("dollar_value_per_referral", 25)
    from .normalizer import normalize_name

    mgr_totals = Counter()
    mgr_self = Counter()
    for rec in records:
        mgr = normalize_name(rec.get("manager"))
        ref = normalize_name(rec.get("referrer"))
        if mgr:
            mgr_totals[mgr] += 1
            if mgr == ref:
                mgr_self[mgr] += 1

    headers = ["Manager", "Score", "Tier", "Total Referrals", "Self-Referrals", "Self-Ref Rate", "Rules Hit"]
    _add_header_row(ws, headers)

    mgr_scores = scores.get("manager", {})
    all_mgrs = set(mgr_totals.keys()) | set(mgr_scores.keys())

    sorted_mgrs = sorted(all_mgrs, key=lambda m: -mgr_scores.get(m, {}).get("score", 0))

    for row_idx, mgr in enumerate(sorted_mgrs, 2):
        data = mgr_scores.get(mgr, {"score": 0, "tier": "LOW", "rules_hit": set()})
        total = mgr_totals.get(mgr, 0)
        self_count = mgr_self.get(mgr, 0)
        rate = self_count / total if total > 0 else 0

        ws.cell(row=row_idx, column=1, value=mgr)
        ws.cell(row=row_idx, column=2, value=data.get("score", 0))
        tier_cell = ws.cell(row=row_idx, column=3, value=data.get("tier", "LOW"))
        tier_cell.fill = TIER_FILLS.get(data.get("tier", "LOW"), PatternFill())
        ws.cell(row=row_idx, column=4, value=total)
        ws.cell(row=row_idx, column=5, value=self_count)
        ws.cell(row=row_idx, column=6, value=f"{rate:.0%}")
        ws.cell(row=row_idx, column=7, value=", ".join(sorted(data.get("rules_hit", set()))))

    _auto_width(ws)


def _write_branch_scorecard(ws, scores, config):
    headers = ["Branch", "Score", "Tier", "Rules Hit", "Finding Count"]
    _add_header_row(ws, headers)

    br_scores = scores.get("branch", {})
    for row_idx, (br, data) in enumerate(
        sorted(br_scores.items(), key=lambda x: -x[1]["score"]), 2
    ):
        ws.cell(row=row_idx, column=1, value=br)
        ws.cell(row=row_idx, column=2, value=data["score"])
        tier_cell = ws.cell(row=row_idx, column=3, value=data["tier"])
        tier_cell.fill = TIER_FILLS.get(data["tier"], PatternFill())
        ws.cell(row=row_idx, column=4, value=", ".join(sorted(data["rules_hit"])))
        ws.cell(row=row_idx, column=5, value=len(data["findings"]))

    _auto_width(ws)


def _write_flagged_records(ws, findings):
    headers = ["Row", "Rule", "Severity", "Referrer", "Manager", "Branch", "Description"]
    _add_header_row(ws, headers)

    row_idx = 2
    for f in sorted(findings, key=lambda x: ({"critical": 0, "high": 1, "medium": 2, "low": 3}.get(x.severity, 4))):
        for row_num in f.row_numbers[:5]:
            ws.cell(row=row_idx, column=1, value=row_num)
            ws.cell(row=row_idx, column=2, value=f.rule_name)
            sev_cell = ws.cell(row=row_idx, column=3, value=f.severity.upper())
            sev_cell.fill = SEVERITY_FILLS.get(f.severity, PatternFill())
            ws.cell(row=row_idx, column=4, value=f.referrer)
            ws.cell(row=row_idx, column=5, value=f.manager)
            ws.cell(row=row_idx, column=6, value=f.branch)
            ws.cell(row=row_idx, column=7, value=f.description)
            row_idx += 1

    _auto_width(ws)


def _write_rule_findings(ws, findings, *rule_names):
    filtered = [f for f in findings if f.rule_name in rule_names]
    headers = ["Rule", "Severity", "Referrer", "Manager", "Branch", "Description", "Rows"]
    _add_header_row(ws, headers)

    for row_idx, f in enumerate(
        sorted(filtered, key=lambda x: -len(x.row_numbers)), 2
    ):
        ws.cell(row=row_idx, column=1, value=f.rule_name)
        sev_cell = ws.cell(row=row_idx, column=2, value=f.severity.upper())
        sev_cell.fill = SEVERITY_FILLS.get(f.severity, PatternFill())
        ws.cell(row=row_idx, column=3, value=f.referrer)
        ws.cell(row=row_idx, column=4, value=f.manager)
        ws.cell(row=row_idx, column=5, value=f.branch)
        ws.cell(row=row_idx, column=6, value=f.description)
        ws.cell(row=row_idx, column=7, value=str(f.row_numbers[:10]))

    _auto_width(ws)


def _write_config_tab(ws, config):
    ws.cell(row=1, column=1, value="Configuration Audit Trail").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    row = 4
    ws.cell(row=row, column=1, value="Setting").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Value").font = HEADER_FONT

    row += 1
    ws.cell(row=row, column=1, value="Client")
    ws.cell(row=row, column=2, value=config.get("client", ""))

    row += 1
    ws.cell(row=row, column=1, value="Program")
    ws.cell(row=row, column=2, value=config.get("program", ""))

    row += 1
    ws.cell(row=row, column=1, value="$/Referral")
    ws.cell(row=row, column=2, value=config.get("dollar_value_per_referral", ""))

    row += 2
    ws.cell(row=row, column=1, value="Rule").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Enabled").font = HEADER_FONT
    ws.cell(row=row, column=3, value="Weight").font = HEADER_FONT

    for rule_key, rule_cfg in config.get("rules", {}).items():
        row += 1
        ws.cell(row=row, column=1, value=rule_key)
        ws.cell(row=row, column=2, value=str(rule_cfg.get("enabled", True)))
        ws.cell(row=row, column=3, value=rule_cfg.get("weight", ""))

    _auto_width(ws)
