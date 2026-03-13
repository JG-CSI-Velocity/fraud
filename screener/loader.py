import csv
from pathlib import Path

import openpyxl
import yaml


def load_config(config_path: str) -> dict:
    with open(config_path) as f:
        client_config = yaml.safe_load(f)

    default_path = Path(config_path).parent / "default.yaml"
    if default_path.exists():
        with open(default_path) as f:
            defaults = yaml.safe_load(f)
        merged = {**defaults, **client_config}
        if "rules" in defaults and "rules" not in client_config:
            merged["rules"] = defaults["rules"]
        if "scoring" in defaults and "scoring" not in client_config:
            merged["scoring"] = defaults["scoring"]
    else:
        merged = client_config

    return merged


def load_records(file_path: str, config: dict) -> list[dict]:
    path = Path(file_path)
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _load_excel(path, config)
    elif path.suffix.lower() == ".csv":
        return _load_csv(path, config)
    else:
        raise ValueError(f"Unsupported file type: {path.suffix}")


def _load_excel(path: Path, config: dict) -> list[dict]:
    sheet_name = config.get("sheet")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [cell.value for cell in ws[1]]
    col_map = config.get("columns", {})
    emp_col = config.get("employee_tracking_column")

    col_indices = {}
    for canonical, header_name in col_map.items():
        if isinstance(header_name, int):
            col_indices[canonical] = header_name - 1
        else:
            try:
                col_indices[canonical] = headers.index(header_name)
            except ValueError:
                pass

    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {"_source_row": row_idx}
        for canonical, col_idx in col_indices.items():
            if col_idx < len(row):
                record[canonical] = row[col_idx]
            else:
                record[canonical] = None

        if emp_col and emp_col - 1 < len(row):
            record["employee_tracking"] = row[emp_col - 1]

        if record.get("manager") or record.get("referrer"):
            records.append(record)

    wb.close()
    return records


def _load_csv(path: Path, config: dict) -> list[dict]:
    col_map = config.get("columns", {})

    records = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, start=2):
            record = {"_source_row": row_idx}
            for canonical, header_name in col_map.items():
                record[canonical] = row.get(str(header_name))
            records.append(record)

    return records
